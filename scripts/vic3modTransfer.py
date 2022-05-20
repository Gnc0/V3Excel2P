#配置
from cgi import print_arguments
from importlib.resources import path
from multiprocessing.dummy import Value
import os
import sys
import openpyxl
from openpyxl import Workbook

#提示信息
print("欢迎使用vic3mod工具表转化器v0.31，请配合0.3版工具表使用。否则可能出现未知错误！")

#文件位置
lujin = input("文件位置：")
os.chdir(lujin)  # 修改工作路径

#获取表格
biaoge = input("表格名称（带后缀）：")
workbook = openpyxl.load_workbook(biaoge)
sheetbg = workbook['建筑组']
sheetb = workbook['建筑']
sheetpmg = workbook['生产方式组']
sheetpm = workbook['生产方式']
sheetbuy = workbook['pop需求量表']
sheetgoods = workbook['货物']
sheetneed = workbook['pop需求类型']
sheethelp = workbook['使用说明']

#检测excel版本
cell_help_v = sheethelp.cell(row=1,column=3)
if cell_help_v.value != 'v0.3':
    input('工具表版本不匹配，请按下回车键关闭程序！')
    sys.exit(1)


#获取pmg表格数据并输出
i_pmg = sheetpmg.max_row
j_pmg = sheetpmg.max_column
with open('mod_pmg.txt','a') as file0:
    for i2_pmg in range(1,i_pmg):
        cell_pmg_name = sheetpmg.cell(row=i2_pmg+1,column=1)
        if cell_pmg_name.value != None:
            print(str(cell_pmg_name.value) +'= { \n	prodution_methods = { \n' ,file=file0)
            j_pmg_end = 0
            for j2_pmg in range (3,j_pmg+1):
                j_pmg_end = j_pmg_end + 1
                cell_pmg_pm = sheetpmg.cell(row=i2_pmg+1,column=j2_pmg)
                if cell_pmg_pm.value != None and j_pmg_end != (j_pmg-2):
                    print('		'+cell_pmg_pm.value+'\n',file=file0)
                elif cell_pmg_pm.value == None and j_pmg_end == (j_pmg-2):
                    print('	} \n}',file=file0)
                elif cell_pmg_pm.value != None and j_pmg_end == (j_pmg-2):
                    print('		'+cell_pmg_pm.value+'\n	} \n}',file=file0)
    print('pmg表格转化完成')

#获取pm配置表格数据并输出
i_pm = sheetpm.max_row
j_pm = sheetpm.max_column
with open('mod_pm.txt','a') as file1:
    for i2_pm in range(1,i_pm):
        cell_pm_name = sheetpm.cell(row=i2_pm+1,column=2)
        if cell_pm_name.value != None:
            print(str(cell_pm_name.value) + '= { \n',file=file1)
            for j2_pm in range (4,j_pm+1):
                cell_pm_title = sheetpm.cell(row=1,column=j2_pm)
                cell_pm_title2 = sheetpm.cell(row=1,column=j2_pm+1)
                cell_pm_title3 = sheetpm.cell(row=1,column=j2_pm+2)
                cell_pm_text = sheetpm.cell(row=i2_pm+1,column=j2_pm)
                cell_pm_text2 = sheetpm.cell(row=i2_pm+1,column=j2_pm+1)
                if cell_pm_title.value == '图标文件位置' :
                    print('	texture = '+ str(cell_pm_text.value) + '\n',file=file1)
                elif cell_pm_title.value == '污染量':
                    print('	pollution_generation = '+ str(cell_pm_text.value) + '\n',file=file1)
                elif cell_pm_title.value == '默认':
                    print('	is_default = '+ str(cell_pm_text.value) + '\n	unlocking_technologies = { \n',file=file1)
                elif cell_pm_title.value == '解锁科技' and cell_pm_title2.value == '解锁科技':
                    if cell_pm_text.value != None:
                        print('		'+cell_pm_text.value +'\n',file=file1)
                elif cell_pm_title.value == '解锁科技' and cell_pm_title2.value != '解锁科技':
                    if cell_pm_text.value != None:
                        print('		'+ cell_pm_text.value+'\n	}\n	unlocking_production_methods = { \n',file=file1)
                    elif cell_pm_text.value == None:
                        print('	}\n	unlocking_production_methods = { \n',file=file1)
                elif cell_pm_title.value == '前置生产方式' and cell_pm_title2.value == '前置生产方式':
                    if cell_pm_text.value != None:
                        print('		'+cell_pm_text.value +'\n',file=file1)        
                elif cell_pm_title.value == '前置生产方式' and cell_pm_title2.value != '前置生产方式':
                    if cell_pm_text.value != None:
                        print('		'+ cell_pm_text.value+'\n	}\n	disallowing_laws = { \n',file=file1)
                    elif cell_pm_text.value == None:
                        print('	}\n	disallowing_laws = { \n',file=file1)
                elif cell_pm_title.value == '禁止法律' and cell_pm_title2.value == '禁止法律':
                    if cell_pm_text.value != None:
                        print('		'+cell_pm_text.value +'\n',file=file1)        
                elif cell_pm_title.value == '禁止法律' and cell_pm_title2.value != '禁止法律':
                    if cell_pm_text.value != None:
                        print('		'+ cell_pm_text.value+'\n	}\n	building_modifiers = { \n		workforce_scaled = {',file=file1)
                    elif cell_pm_text.value == None:
                        print('	}\n	building_modifiers = { \n		workforce_scaled = {',file=file1)
                elif cell_pm_title.value == '投入货物' and cell_pm_title3.value == '投入货物':
                    if cell_pm_text.value != None:
                        print('		building_input_'+cell_pm_text.value+'_add = '+str(cell_pm_text2.value)+'\n',file=file1)        
                elif cell_pm_title.value == '投入货物' and cell_pm_title3.value != '投入货物':
                    if cell_pm_text.value != None:
                        print('		building_input_'+cell_pm_text.value+'_add = '+str(cell_pm_text2.value)+'\n		}\n		level_scaled = { \n',file=file1)
                    elif cell_pm_text.value == None:
                        print('		}\n	level_scaled = { \n',file=file1)        
                elif cell_pm_title.value == '雇佣pop' and cell_pm_title3.value == '雇佣pop':
                    if cell_pm_text.value != None:
                        print('			building_employment_'+cell_pm_text.value+'_add = '+str(cell_pm_text2.value)+'\n',file=file1)        
                elif cell_pm_title.value == '雇佣pop' and cell_pm_title3.value != '雇佣pop':
                    if cell_pm_text.value != None:
                        print('			building_employment_'+cell_pm_text.value+'_add = '+str(cell_pm_text2.value)+'\n',file=file1)
                elif cell_pm_title.value == '成长效果' and cell_pm_title3.value == '成长效果':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n',file=file1)        
                elif cell_pm_title.value == '成长效果' and cell_pm_title3.value != '成长效果':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n		}\n		unscaled = { \n',file=file1)
                    elif cell_pm_text.value == None:
                        print('		}\n		unscaled = { \n',file=file1)  
                elif cell_pm_title.value == '不成长效果' and cell_pm_title3.value == '不成长效果':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n',file=file1)        
                elif cell_pm_title.value == '不成长效果' and cell_pm_title3.value != '不成长效果':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n		}\n	}\n	country_modifiers = { \n		workforce_scaled = { \n',file=file1)
                    elif cell_pm_text.value == None:
                        print('		}\n	}\n	country_modifiers = { \n		workforce_scaled = { \n',file=file1)  
                elif cell_pm_title.value == '国家修正' and cell_pm_title3.value == '国家修正':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n',file=file1)        
                elif cell_pm_title.value == '国家修正' and cell_pm_title3.value != '国家修正':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n		}\n	}\n	timed_modifiers = { \n',file=file1)
                    elif cell_pm_text.value == None:
                        print('		}\n	}\n	timed_modifiers = { \n',file=file1) 
                elif cell_pm_title.value == '限时修正' and cell_pm_title3.value == '限时修正':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n',file=file1)        
                elif cell_pm_title.value == '限时修正' and cell_pm_title3.value != '限时修正':
                    if cell_pm_text.value != None:
                        print('			'+cell_pm_text.value+' = '+str(cell_pm_text2.value)+'\n		}\n	}\n',file=file1)
                    elif cell_pm_text.value == None:
                        print('		}\n	}\n',file=file1) 
    print('pm表格转化完成')

#获取buy表数据并输出
i_buy = sheetbuy.max_row
j_buy = sheetbuy.max_column
with open('mod_buy.txt','a') as file2:
    for i2_buy in range(8,i_buy):
        cell_buy_wealth = sheetbuy.cell(row=i2_buy,column=1)
        cell_buy_political = sheetbuy.cell(row=i2_buy,column=2)
        print('wealth_'+str(cell_buy_wealth.value) + ' = { \n	political_strength = '+str(cell_buy_political.value)+'\n	goods = { \n',file=file2)
        for j2_buy in range (4,j_buy):   
            cell_buy_name = sheetbuy.cell(row=2,column=j2_buy)
            cell_buy_name2 = sheetbuy.cell(row=2,column=j2_buy+1)            
            cell_buy_text = sheetbuy.cell(row=i2_buy,column=j2_buy)
            if cell_buy_name.value != None and cell_buy_name2.value != None:
                if cell_buy_text.value != None:
                    print('		popneed_'+cell_buy_name.value+' = '+str(cell_buy_text.value)+'\n',file=file2)
            if cell_buy_name.value != None and cell_buy_name2.value == None:
                if cell_buy_text.value != None:
                    print('		popneed_'+cell_buy_name.value+' = '+str(cell_buy_text.value)+'\n	} \n}',file=file2)
                if cell_buy_text.value == None:            
                    print('	} \n}',file=file2)    
    print('buy表格转化完成')    

#获取goods表数据并输出
i_goods = sheetgoods.max_row
j_goods = sheetgoods.max_column
with open('mod_goods.txt','a') as file3:
    for i2_goods in range(2,i_goods+1):
        cell_goods_name = sheetgoods.cell(row=i2_goods,column=1)
        print(cell_goods_name.value+' = { \n',file=file3)      
        for j2_goods in range (3,j_goods+1):
            cell_goods_text = sheetgoods.cell(row=i2_goods,column=j2_goods)
            if j2_goods == 3 and cell_goods_text.value != None:
                print('	texture = '+cell_goods_text.value+'\n',file=file3)
            if j2_goods == 4 and cell_goods_text.value != None:
                print('	cost = '+str(cell_goods_text.value)+'\n',file=file3)
            if j2_goods == 5 and cell_goods_text.value != None:
                print('	category = '+cell_goods_text.value+'\n',file=file3)    
            if j2_goods == 6 and cell_goods_text.value != None:
                print('	obsession_chance = '+str(cell_goods_text.value)+'\n',file=file3)
            if j2_goods == 7 and cell_goods_text.value != None:
                print('	prestige_factor = '+str(cell_goods_text.value)+'\n',file=file3)
            if j2_goods == 8 and cell_goods_text.value != None:
                print('	tradeable = '+cell_goods_text.value+'\n',file=file3)  
            if j2_goods == 9 and cell_goods_text.value != None:
                print('	trade_route_revenue_multiplier = '+str(cell_goods_text.value)+'\n',file=file3)
            if j2_goods == 10 and cell_goods_text.value != None:
                print('	convoy_shipping_multiplier = '+str(cell_goods_text.value)+'\n}',file=file3)
            if j2_goods == 10 and cell_goods_text.value == None:  
                print('} \n',file=file3)              
    print('goods表格转化完成')               

#获取need表数据并输出
i_need = sheetneed.max_row
j_need = sheetneed.max_column
with open('mod_need.txt','a') as file4:
    for i2_need in range(1,i_need):    
        cell_need_name = sheetneed.cell(row=i2_need+1,column=1)
        cell_need_default = sheetneed.cell(row=i2_need+1,column=3)
        if cell_need_name.value != None:
            print('popneed_'+str(cell_need_name.value) + ' = { \n',file=file4)
            print('	default = '+str(cell_need_default.value) + '\n',file=file4)
            for j2_need in range (3,j_need+2):
                cell_need_title = sheetneed.cell(row=1,column=j2_need)            
                cell_need_text = sheetneed.cell(row=i2_need+1,column=j2_need)
                cell_need_text2 = sheetneed.cell(row=i2_need+1,column=j2_need+1)
                cell_need_text3 = sheetneed.cell(row=i2_need+1,column=j2_need+2)
                cell_need_text4 = sheetneed.cell(row=i2_need+1,column=j2_need+3)
                if cell_need_title.value == '替代品' and cell_need_text.value != None:
                    print('	entry = { \n		goods = '+str(cell_need_text.value)+'\n',file=file4)
                    print('		weight = '+str(cell_need_text2.value)+'\n',file=file4)
                    print('		max_weight = '+str(cell_need_text3.value)+'\n',file=file4)
                    print('		min_weight = '+str(cell_need_text4.value)+'\n	}\n',file=file4)
                if j2_need == (j_need+1):
                    print('}',file=file4)
    print('need表格转化完成')

input("转化完成，按下回车键退出")